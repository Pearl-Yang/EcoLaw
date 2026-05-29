"""
AI内容分析服务 - 使用AI总结帖子和视频内容
支持本地分析和调用大模型API两种模式
"""

import re
import os
from typing import List, Dict, Any, Optional
from dataclasses import dataclass, field


@dataclass
class ContentSummary:
    """内容摘要"""
    content_id: str = ""           # 内容ID
    platform: str = ""             # 平台
    original_title: str = ""       # 原标题
    original_content: str = ""     # 原文内容
    summary: str = ""              # AI摘要
    key_points: List[str] = field(default_factory=list)  # 关键点
    public_opinion: str = ""       # 舆论倾向
    tags: List[str] = field(default_factory=list)  # 标签
    word_count: int = 0            # 字数
    summary_by: str = "规则"       # 总结方式


class AIAnalyzer:
    """AI内容分析器"""

    def __init__(self, api_key: Optional[str] = None, model: str = "gpt-3.5-turbo"):
        """
        初始化AI分析器

        Args:
            api_key: OpenAI API密钥（可选，不提供则使用本地规则分析）
            model: 使用的模型
        """
        self.api_key = api_key or os.getenv("OPENAI_API_KEY")
        self.model = model
        self.use_remote = bool(self.api_key)

    def summarize_posts(self, posts: List[Dict[str, Any]], batch_size: int = 10) -> List[ContentSummary]:
        """批量总结帖子内容"""
        summaries = []

        for i, post in enumerate(posts):
            try:
                summary = self._summarize_single_post(post)
                summaries.append(summary)

                if (i + 1) % batch_size == 0:
                    print(f"  已处理 {i + 1}/{len(posts)} 个帖子")

            except Exception as e:
                print(f"  总结帖子失败: {e}")
                # 返回基本摘要
                summaries.append(ContentSummary(
                    content_id=str(post.get("post_id", "")),
                    platform=post.get("platform", ""),
                    original_title=post.get("title", ""),
                    original_content=post.get("content", ""),
                    summary="总结失败",
                    summary_by="错误"
                ))

        return summaries

    def summarize_videos(self, videos: List[Dict[str, Any]], batch_size: int = 10) -> List[ContentSummary]:
        """批量总结视频内容"""
        summaries = []

        for i, video in enumerate(videos):
            try:
                summary = self._summarize_single_video(video)
                summaries.append(summary)

                if (i + 1) % batch_size == 0:
                    print(f"  已处理 {i + 1}/{len(videos)} 个视频")

            except Exception as e:
                print(f"  总结视频失败: {e}")
                summaries.append(ContentSummary(
                    content_id=str(video.get("video_id", "")),
                    platform=video.get("platform", ""),
                    original_title=video.get("title", ""),
                    original_content=video.get("description", ""),
                    summary="总结失败",
                    summary_by="错误"
                ))

        return summaries

    def _summarize_single_post(self, post: Dict[str, Any]) -> ContentSummary:
        """总结单个帖子"""
        title = post.get("title", "") or ""
        content = post.get("content", "") or ""
        full_text = f"{title} {content}".strip()

        if self.use_remote:
            summary_text = self._get_remote_summary(full_text)
        else:
            summary_text = self._get_local_summary(full_text)

        key_points = self._extract_key_points(full_text)
        tags = self._extract_tags(full_text)
        opinion = self._analyze_public_opinion(full_text)

        return ContentSummary(
            content_id=str(post.get("post_id", "")),
            platform=post.get("platform", ""),
            original_title=title,
            original_content=content,
            summary=summary_text,
            key_points=key_points,
            public_opinion=opinion,
            tags=tags,
            word_count=len(full_text),
            summary_by="AI" if self.use_remote else "规则"
        )

    def _summarize_single_video(self, video: Dict[str, Any]) -> ContentSummary:
        """总结单个视频"""
        title = video.get("title", "") or ""
        description = video.get("description", "") or ""
        full_text = f"{title} {description}".strip()

        if self.use_remote:
            summary_text = self._get_remote_summary(full_text)
        else:
            summary_text = self._get_local_summary(full_text)

        key_points = self._extract_key_points(full_text)
        tags = self._extract_tags(full_text)
        opinion = self._analyze_public_opinion(full_text)

        return ContentSummary(
            content_id=str(video.get("video_id", "")),
            platform=video.get("platform", ""),
            original_title=title,
            original_content=description,
            summary=summary_text,
            key_points=key_points,
            public_opinion=opinion,
            tags=tags,
            word_count=len(full_text),
            summary_by="AI" if self.use_remote else "规则"
        )

    def _get_remote_summary(self, text: str) -> str:
        """调用远程API获取摘要"""
        try:
            import openai

            client = openai.OpenAI(api_key=self.api_key)

            response = client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": "你是一个内容分析助手。请用50字以内总结以下内容的核心观点。"},
                    {"role": "user", "content": text[:2000]}
                ],
                max_tokens=100,
                temperature=0.7
            )

            return response.choices[0].message.content.strip()

        except Exception as e:
            print(f"  API调用失败，回退到本地分析: {e}")
            return self._get_local_summary(text)

    def _get_local_summary(self, text: str) -> str:
        """本地规则生成摘要"""
        if not text:
            return "内容为空"

        # 提取关键句子
        sentences = re.split(r'[。！？\n]', text)
        sentences = [s.strip() for s in sentences if s.strip()]

        if not sentences:
            return text[:100] + "..." if len(text) > 100 else text

        # 选择最长的句子作为摘要
        longest_sentence = max(sentences, key=len)
        summary = longest_sentence[:150] + "..." if len(longest_sentence) > 150 else longest_sentence

        return summary

    def _extract_key_points(self, text: str) -> List[str]:
        """提取关键点"""
        key_points = []

        # 关键词模式
        patterns = [
            r'([\u4e00-\u9fa5]{2,}政策[\u4e00-\u9fa5]*)',
            r'([\u4e00-\u9fa5]{2,}措施[\u4e00-\u9fa5]*)',
            r'([\u4e00-\u9fa5]{2,}规定[\u4e00-\u9fa5]*)',
            r'需要([\u4e00-\u9fa5]+)',
            r'应该([\u4e00-\u9fa5]+)',
            r'建议([\u4e00-\u9fa5]+)',
        ]

        for pattern in patterns:
            matches = re.findall(pattern, text)
            for match in matches[:2]:  # 每个模式最多取2个
                if match not in key_points and len(match) > 2:
                    key_points.append(match)

        # 如果没找到，返回前几个关键词
        if not key_points:
            words = re.findall(r'[\u4e00-\u9fa5]{3,5}', text)
            from collections import Counter
            word_counts = Counter(words)
            key_points = [w for w, _ in word_counts.most_common(3)]

        return key_points[:5]

    def _extract_tags(self, text: str) -> List[str]:
        """提取标签"""
        tags = []

        # 主题标签
        topic_keywords = {
            "智慧普法": ["智慧普法", "智能普法", "数字普法", "普法信息化", "智慧司法", "普法科技", "数字化普法", "智慧法治", "普法平台"],
            "生态环境法": ["生态环境法典", "环境法典", "环保法典", "生态环境法律", "环境法律体系", "环保法律法规", "生态环境法治", "环境法治建设"],
            "环保立法": ["生态环境保护法", "环保立法", "环境法典化", "生态环境法律规定"],
            "政策": ["政策", "规定", "条例", "法规", "法律"],
            "公众参与": ["公众", "群众", "百姓", "民众", "社会", "人人"],
            "普法教育": ["普法", "宣传教育", "法治教育", "法律知识"],
        }

        for tag, keywords in topic_keywords.items():
            if any(kw in text for kw in keywords):
                tags.append(tag)

        return list(set(tags))[:5]  # 去重并限制数量

    def _analyze_public_opinion(self, text: str) -> str:
        """分析舆论倾向"""
        # 智慧普法和生态环境法律相关的情感词汇
        positive_words = [
            "支持", "点赞", "好", "棒", "赞", "应该", "重要", "必要", "积极",
            "有用", "实用", "便捷", "方便", "高效", "智慧", "科学", "规范",
            "保护", "维护", "改善", "提升", "推进", "促进"
        ]
        negative_words = [
            "反对", "吐槽", "差", "垃圾", "坑", "骗", "假", "没用", "无用",
            "麻烦", "不便", "困难", "难", "贵", "形式主义", "��话", "脱离实际",
            "担忧", "担心", "质疑", "疑问", "不理解", "不接受"
        ]

        positive_count = sum(1 for w in positive_words if w in text)
        negative_count = sum(1 for w in negative_words if w in text)

        if positive_count > negative_count:
            return "正面"
        elif negative_count > positive_count:
            return "负面"
        else:
            return "中性"


class SummaryExporter:
    """摘要导出器"""

    @staticmethod
    def export_summary_to_excel(summaries: List[ContentSummary], filename: str):
        """导出摘要到Excel"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "内容摘要"

        # 表头
        headers = ["平台", "ID", "原标题", "内容摘要", "关键点", "舆论倾向", "标签", "字数", "总结方式"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 数据行
        for row_idx, summary in enumerate(summaries, 2):
            ws.cell(row=row_idx, column=1, value=summary.platform)
            ws.cell(row=row_idx, column=2, value=summary.content_id)
            ws.cell(row=row_idx, column=3, value=summary.original_title)
            ws.cell(row=row_idx, column=4, value=summary.summary)
            ws.cell(row=row_idx, column=5, value="\n".join(summary.key_points))
            ws.cell(row=row_idx, column=6, value=summary.public_opinion)
            ws.cell(row=row_idx, column=7, value=", ".join(summary.tags))
            ws.cell(row=row_idx, column=8, value=summary.word_count)
            ws.cell(row=row_idx, column=9, value=summary.summary_by)

        # 设置列宽
        widths = [12, 15, 30, 50, 40, 10, 20, 10, 12]
        from openpyxl.utils import get_column_letter
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = "A2"
        wb.save(filename)
        return filename


def main():
    """测试"""
    import json

    # 读取数据
    with open("../data/crawl_results_20260404_171001.json", "r", encoding="utf-8") as f:
        data = json.load(f)

    result = data[0]
    posts = result.get("posts", [])
    videos = result.get("videos", [])

    print(f"帖子数: {len(posts)}, 视频数: {len(videos)}")

    # 分析
    analyzer = AIAnalyzer()
    post_summaries = analyzer.summarize_posts(posts[:5])  # 测试前5个
    video_summaries = analyzer.summarize_videos(videos[:5])

    print("\n帖子摘要示例:")
    for s in post_summaries[:2]:
        print(f"  标题: {s.original_title[:30]}...")
        print(f"  摘要: {s.summary}")
        print(f"  关键点: {s.key_points}")
        print(f"  舆论: {s.public_opinion}")
        print()

    print("\n视频摘要示例:")
    for s in video_summaries[:2]:
        print(f"  标题: {s.original_title[:30]}...")
        print(f"  摘要: {s.summary}")
        print(f"  关键点: {s.key_points}")
        print(f"  舆论: {s.public_opinion}")
        print()

    # 导出
    all_summaries = post_summaries + video_summaries
    exporter = SummaryExporter()
    filename = exporter.export_summary_to_excel(all_summaries, "../data/content_summaries.xlsx")
    print(f"摘要已导出: {filename}")


if __name__ == "__main__":
    main()
