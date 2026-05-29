"""
Excel导出器 - 将爬取的数据导出为Excel文件
"""

import os
import json
import ast
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Excel导出器"""

    def __init__(self, output_dir: str = "data"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

        # 样式定义
        self.header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        self.cell_font = Font(name="微软雅黑", size=10)
        self.cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        self.positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.neutral_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    def _to_dict(self, item) -> Dict:
        """将item转换为字典"""
        if isinstance(item, dict):
            return item
        if hasattr(item, 'to_dict'):
            return item.to_dict()
        if isinstance(item, str):
            # 尝试解析字符串表示
            try:
                return ast.literal_eval(item)
            except:
                return {}
        return {}

    def _get_attr(self, item, attr, default=None):
        """安全获取属性"""
        if isinstance(item, dict):
            return item.get(attr, default)
        if hasattr(item, attr):
            return getattr(item, attr)
        return default

    def _apply_header_style(self, cell):
        """应用表头样式"""
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.header_alignment
        cell.border = self.border

    def _apply_cell_style(self, cell, sentiment: str = None):
        """应用单元格样式"""
        cell.font = self.cell_font
        cell.alignment = self.cell_alignment
        cell.border = self.border

        if sentiment:
            if sentiment == "正面":
                cell.fill = self.positive_fill
            elif sentiment == "负面":
                cell.fill = self.negative_fill
            else:
                cell.fill = self.neutral_fill

    def _set_column_width(self, worksheet, widths: List[int]):
        """设置列宽"""
        for i, width in enumerate(widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width

    def export_posts(self, posts: List[Any], filename: str = None) -> str:
        """导出帖子数据到Excel"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = os.path.join(self.output_dir, f"posts_{timestamp}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "帖子数据"

        # 表头
        headers = [
            "平台", "帖子ID", "标题", "内容", "作者", "作者ID",
            "发布时间", "点赞数", "评论数", "转发数", "浏览量",
            "链接", "标签", "情感倾向", "匹配关键词"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # 数据行
        for row_idx, post in enumerate(posts, 2):
            values = self._to_dict(post) if not isinstance(post, dict) else post
            for col, key in enumerate(headers, 1):
                value = self._get_attr(values, key, "")
                cell = ws.cell(row=row_idx, column=col, value=value)
                sentiment = self._get_attr(values, "情感倾向", "")
                self._apply_cell_style(cell, sentiment)

        # 设置列宽
        self._set_column_width(ws, [10, 15, 25, 40, 12, 15, 18, 10, 10, 10, 10, 30, 20, 10, 20])

        # 冻结首行
        ws.freeze_panes = "A2"

        wb.save(filename)
        return filename

    def export_videos(self, videos: List[Any], filename: str = None) -> str:
        """导出视频数据到Excel"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = os.path.join(self.output_dir, f"videos_{timestamp}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "视频数据"

        # 表头（与VideoItem字段对应）
        headers = [
            "平台", "视频ID", "标题", "描述", "作者", "作者ID",
            "发布时间", "时长(秒)", "点赞数", "评论数", "转发数",
            "播放量", "投币数", "收藏数", "弹幕数",
            "链接", "封面链接", "标签", "分类", "情感倾向", "匹配关键词"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # 数据行
        for row_idx, video in enumerate(videos, 2):
            values = self._to_dict(video) if not isinstance(video, dict) else video
            for col, key in enumerate(headers, 1):
                # 处理特殊字段
                if key == "时长(秒)":
                    value = self._get_attr(values, "duration", 0)
                elif key == "封面链接":
                    value = self._get_attr(values, "cover_url", "")
                elif key == "匹配关键词":
                    # 将关键词列表转为字符串
                    keywords = self._get_attr(values, "keywords", [])
                    value = ",".join(keywords) if isinstance(keywords, list) else str(keywords)
                elif key == "标签":
                    # 将标签列表转为字符串
                    tags = self._get_attr(values, "tags", [])
                    value = ",".join(tags) if isinstance(tags, list) else str(tags)
                else:
                    value = self._get_attr(values, key, "")
                cell = ws.cell(row=row_idx, column=col, value=value)
                sentiment = self._get_attr(values, "情感倾向", "")
                self._apply_cell_style(cell, sentiment)

        # 设置列宽
        self._set_column_width(ws, [10, 15, 35, 40, 12, 15, 18, 10, 10, 10, 10, 12, 10, 10, 10, 35, 35, 20, 10, 10, 20])

        # 冻结首行
        ws.freeze_panes = "A2"

        wb.save(filename)
        return filename

    def export_all(self, crawler_results: List[Dict[str, Any]], filename: str = None) -> str:
        """导出所有数据到一个Excel文件（多个Sheet）"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = os.path.join(self.output_dir, f"social_data_{timestamp}.xlsx")

        wb = Workbook()

        # 删除默认sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # 收集所有帖子和视频
        all_posts = []
        all_videos = []

        for result in crawler_results:
            all_posts.extend(result.get("posts", []))
            all_videos.extend(result.get("videos", []))

        # 创建汇总统计sheet
        ws_summary = wb.create_sheet("汇总统计")
        self._create_summary_sheet(ws_summary, crawler_results)

        # 创建帖子sheet
        if all_posts:
            ws_posts = wb.create_sheet("帖子数据")
            self._create_posts_sheet(ws_posts, all_posts)

        # 创建视频sheet
        if all_videos:
            ws_videos = wb.create_sheet("视频数据")
            self._create_videos_sheet(ws_videos, all_videos)

        # 按平台创建sheet
        platforms = set(r.get("platform", "") for r in crawler_results)
        for platform in platforms:
            platform_posts = [p for r in crawler_results if r.get("platform") == platform for p in r.get("posts", [])]
            platform_videos = [v for r in crawler_results if r.get("platform") == platform for v in r.get("videos", [])]

            if platform_posts or platform_videos:
                sheet_name = platform[:31]  # Excel sheet名最多31字符
                ws_platform = wb.create_sheet(sheet_name)

                if platform_posts:
                    ws_platform.title = f"{platform[:25]}帖子"
                    self._create_posts_sheet(ws_platform, platform_posts)
                elif platform_videos:
                    ws_platform.title = f"{platform[:25]}视频"
                    self._create_videos_sheet(ws_platform, platform_videos)

        wb.save(filename)
        return filename

    def _create_summary_sheet(self, ws, crawler_results: List[Dict[str, Any]]):
        """创建汇总统计sheet"""
        # 标题
        ws.merge_cells("A1:E1")
        title_cell = ws.cell(row=1, column=1, value="白色污染治理舆情数据汇总")
        title_cell.font = Font(name="微软雅黑", size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        # 爬取时间
        ws.cell(row=2, column=1, value="爬取时间").font = Font(bold=True)
        ws.cell(row=2, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        # 统计表头
        row = 4
        headers = ["平台", "帖子数量", "视频数量", "总互动数", "正面占比"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_header_style(cell)

        # 统计数据
        row = 5
        for result in crawler_results:
            platform = result.get("platform", "")
            posts = result.get("posts", [])
            videos = result.get("videos", [])

            # 计算总互动数
            total_interactions = 0
            for p in posts:
                likes = p.likes if hasattr(p, 'likes') else (p.get('likes', 0) if isinstance(p, dict) else 0)
                comments = p.comments if hasattr(p, 'comments') else (p.get('comments', 0) if isinstance(p, dict) else 0)
                shares = p.shares if hasattr(p, 'shares') else (p.get('shares', 0) if isinstance(p, dict) else 0)
                total_interactions += likes + comments + shares

            for v in videos:
                likes = v.likes if hasattr(v, 'likes') else (v.get('likes', 0) if isinstance(v, dict) else 0)
                comments = v.comments if hasattr(v, 'comments') else (v.get('comments', 0) if isinstance(v, dict) else 0)
                shares = v.shares if hasattr(v, 'shares') else (v.get('shares', 0) if isinstance(v, dict) else 0)
                total_interactions += likes + comments + shares

            # 统计正面数量
            positive_count = 0
            for p in posts:
                sentiment = p.sentiment if hasattr(p, 'sentiment') else (p.get('sentiment', '') if isinstance(p, dict) else '')
                if sentiment == "正面":
                    positive_count += 1

            for v in videos:
                sentiment = v.sentiment if hasattr(v, 'sentiment') else (v.get('sentiment', '') if isinstance(v, dict) else '')
                if sentiment == "正面":
                    positive_count += 1

            total_count = len(posts) + len(videos)
            positive_ratio = f"{positive_count / total_count * 100:.1f}%" if total_count > 0 else "0%"

            ws.cell(row=row, column=1, value=platform)
            ws.cell(row=row, column=2, value=len(posts))
            ws.cell(row=row, column=3, value=len(videos))
            ws.cell(row=row, column=4, value=total_interactions)
            ws.cell(row=row, column=5, value=positive_ratio)

            for col in range(1, 6):
                self._apply_cell_style(ws.cell(row=row, column=col))

            row += 1

        # 总计
        row += 1
        ws.cell(row=row, column=1, value="总计").font = Font(bold=True)
        total_posts = sum(len(r.get("posts", [])) for r in crawler_results)
        total_videos = sum(len(r.get("videos", [])) for r in crawler_results)
        ws.cell(row=row, column=2, value=total_posts)
        ws.cell(row=row, column=3, value=total_videos)

        # 设置列宽
        self._set_column_width(ws, [15, 12, 12, 15, 12])

    def _create_posts_sheet(self, ws, posts: List[Any]):
        """创建帖子数据sheet"""
        headers = [
            "平台", "帖子ID", "标题", "内容", "作者", "作者ID",
            "发布时间", "点赞数", "评论数", "转发数", "浏览量",
            "链接", "标签", "情感倾向", "匹配关键词"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        for row_idx, post in enumerate(posts, 2):
            values = self._to_dict(post) if not isinstance(post, dict) else post
            for col, key in enumerate(headers, 1):
                value = self._get_attr(values, key, "")
                cell = ws.cell(row=row_idx, column=col, value=value)
                sentiment = self._get_attr(values, "情感倾向", "")
                self._apply_cell_style(cell, sentiment)

        self._set_column_width(ws, [10, 15, 25, 40, 12, 15, 18, 10, 10, 10, 10, 30, 20, 10, 20])
        ws.freeze_panes = "A2"

    def _create_videos_sheet(self, ws, videos: List[Any]):
        """创建视频数据sheet"""
        headers = [
            "平台", "视频ID", "标题", "描述", "作者", "作者ID",
            "发布时间", "时长(秒)", "点赞数", "评论数", "转发数",
            "播放量", "投币数", "收藏数", "弹幕数",
            "链接", "封面链接", "标签", "分类", "情感倾向", "匹配关键词"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        for row_idx, video in enumerate(videos, 2):
            values = self._to_dict(video) if not isinstance(video, dict) else video
            for col, key in enumerate(headers, 1):
                # 处理特殊字段
                if key == "时长(秒)":
                    value = self._get_attr(values, "duration", 0)
                elif key == "封面链接":
                    value = self._get_attr(values, "cover_url", "")
                elif key == "匹配关键词":
                    keywords = self._get_attr(values, "keywords", [])
                    value = ",".join(keywords) if isinstance(keywords, list) else str(keywords)
                elif key == "标签":
                    tags = self._get_attr(values, "tags", [])
                    value = ",".join(tags) if isinstance(tags, list) else str(tags)
                else:
                    value = self._get_attr(values, key, "")
                cell = ws.cell(row=row_idx, column=col, value=value)
                sentiment = self._get_attr(values, "情感倾向", "")
                self._apply_cell_style(cell, sentiment)

        self._set_column_width(ws, [10, 15, 35, 40, 12, 15, 18, 10, 10, 10, 10, 12, 10, 10, 10, 35, 35, 20, 10, 10, 20])
        ws.freeze_panes = "A2"


def export_to_json(data: Any, filename: str) -> str:
    """导出数据为JSON格式"""
    with open(filename, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=str)
    return filename


if __name__ == "__main__":
    # 测试
    from base_crawler import PostItem, VideoItem

    posts = [
        PostItem(
            platform="哔哩哔哩",
            post_id="123",
            title="测试帖子",
            content="这是一条测试内容",
            author="测试用户",
            likes=100,
            comments=20,
            sentiment="正面"
        )
    ]

    exporter = ExcelExporter()
    filename = exporter.export_posts(posts)
    print(f"导出完成: {filename}")
