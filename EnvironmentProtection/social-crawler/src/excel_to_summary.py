"""从Excel读取数据并生成AI摘要"""
import sys
sys.path.insert(0, '.')
import openpyxl
from ai_analyzer import AIAnalyzer, SummaryExporter, ContentSummary

def read_excel_data(filename):
    """从Excel读取帖子和视频数据"""
    wb = openpyxl.load_workbook(filename)
    
    posts = []
    videos = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            continue
            
        # 读取表头
        headers = [ws.cell(1, i).value for i in range(1, ws.max_column + 1)]
        
        # 读取数据行
        for row_idx in range(2, ws.max_row + 1):
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                row_data[header] = ws.cell(row_idx, col_idx).value
            
            # 根据sheet名称分类
            if '帖子' in sheet_name:
                posts.append(row_data)
            elif '视频' in sheet_name:
                videos.append(row_data)
    
    return posts, videos

def main():
    # 读取Excel数据
    print('正在读取Excel数据...')
    posts, videos = read_excel_data('../data/social_data_20260404_172648.xlsx')
    
    print(f'读取到 {len(posts)} 个帖子, {len(videos)} 个视频')
    
    # 显示示例
    if posts:
        p = posts[0]
        print(f'\n第一个帖子:')
        print(f'  平台: {p.get("平台")}')
        print(f'  帖子ID: {p.get("帖子ID")}')
        print(f'  标题: {p.get("标题")}')
        print(f'  内容: {str(p.get("内容", ""))[:100]}...')
    
    if videos:
        v = videos[0]
        print(f'\n第一个视频:')
        print(f'  平台: {v.get("平台")}')
        print(f'  视频ID: {v.get("视频ID")}')
        print(f'  标题: {v.get("标题")}')
        print(f'  描述: {str(v.get("描述", ""))[:100]}...')
        print(f'  时长: {v.get("时长(秒)")}秒')
        print(f'  播放量: {v.get("播放量")}')
        print(f'  点赞: {v.get("点赞数")}')
    
    # 分析
    print('\n正在分析帖子...')
    analyzer = AIAnalyzer()
    
    # 准备帖子摘要数据
    post_summaries = []
    for p in posts[:20]:  # 分析前20个
        summary = ContentSummary(
            content_id=str(p.get('帖子ID', '')),
            platform=str(p.get('平台', '')),
            original_title=str(p.get('标题', '')),
            original_content=str(p.get('内容', '')),
            summary=analyzer._get_local_summary(str(p.get('标题', '')) + ' ' + str(p.get('内容', ''))),
            key_points=analyzer._extract_key_points(str(p.get('标题', '')) + ' ' + str(p.get('内容', ''))),
            public_opinion=analyzer._analyze_public_opinion(str(p.get('标题', '')) + ' ' + str(p.get('内容', ''))),
            tags=analyzer._extract_tags(str(p.get('标题', '')) + ' ' + str(p.get('内容', ''))),
            word_count=len(str(p.get('内容', ''))),
            summary_by='规则'
        )
        post_summaries.append(summary)
        if len(post_summaries) % 5 == 0:
            print(f'  已处理 {len(post_summaries)}/{min(len(posts), 20)}')
    
    # 准备视频摘要数据
    print('\n正在分析视频...')
    video_summaries = []
    for v in videos[:20]:  # 分析前20个
        title = str(v.get('标题', ''))
        desc = str(v.get('描述', ''))
        full_text = title + ' ' + desc
        
        summary = ContentSummary(
            content_id=str(v.get('视频ID', '')),
            platform=str(v.get('平台', '')),
            original_title=title,
            original_content=desc,
            summary=analyzer._get_local_summary(full_text),
            key_points=analyzer._extract_key_points(full_text),
            public_opinion=analyzer._analyze_public_opinion(full_text),
            tags=analyzer._extract_tags(full_text),
            word_count=len(desc),
            summary_by='规则'
        )
        video_summaries.append(summary)
        if len(video_summaries) % 5 == 0:
            print(f'  已处理 {len(video_summaries)}/{min(len(videos), 20)}')
    
    # 导出
    all_summaries = post_summaries + video_summaries
    exporter = SummaryExporter()
    filename = exporter.export_summary_to_excel(all_summaries, '../data/content_summaries_final.xlsx')
    print(f'\n摘要已导出: {filename}')
    
    # 显示示例
    print('\n帖子摘要示例:')
    for s in post_summaries[:3]:
        print(f'  标题: {s.original_title[:40]}...' if len(s.original_title) > 40 else f'  标题: {s.original_title}')
        print(f'  摘要: {s.summary}')
        print(f'  关键点: {s.key_points}')
        print(f'  舆论: {s.public_opinion}')
        print()
    
    print('视频摘要示例:')
    for s in video_summaries[:3]:
        print(f'  标题: {s.original_title[:40]}...' if len(s.original_title) > 40 else f'  标题: {s.original_title}')
        print(f'  摘要: {s.summary}')
        print(f'  关键点: {s.key_points}')
        print(f'  舆论: {s.public_opinion}')
        print()

if __name__ == '__main__':
    main()
